from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from .models import LeaderboardPayload, RankingEntry

HEADER_ALIASES: Dict[str, List[str]] = {
    "rank": ["rank", "position", "#"],
    "school_name": ["school", "institution", "university", "business school"],
    "program_name": ["program", "programme", "degree", "course", "programme name"],
    "country": ["country", "location"],
    "city": ["city", "campus"],
    "score": ["score", "points", "index"],
    "link": ["link", "url", "website"],
}

FILENAME_PATTERNS: Sequence[Tuple[str, str, str]] = [
    ("masters-in-management", "mim", "Master in Management"),
    ("masters-of-management", "mim", "Master in Management"),
    ("masters-in-finance-pre-experience", "finance", "Master in Finance (pre-experience)"),
    ("masters-in-finance-post-experience", "finance", "Master in Finance (post-experience)"),
    ("masters-in-finance", "finance", "Master in Finance"),
    ("european-business-school-rankings", "business_school", "European Business School Rankings"),
    ("executive-education-custom", "executive_education", "Executive Education Custom"),
    ("executive-education-open", "executive_education", "Executive Education Open"),
    ("executive-mba", "emba", "Executive MBA"),
    ("emba", "emba", "Executive MBA"),
    ("online-mba", "mba", "Online MBA"),
    ("mba", "mba", "MBA"),
]

URL_MAP: Dict[Tuple[str, int], str] = {
    ("masters-in-management", 2021): "https://rankings.ft.com/rankings/5/masters-in-management-2021",
    ("masters-in-management", 2022): "https://rankings.ft.com/rankings/2875/masters-of-management-2022",
    ("masters-in-management", 2023): "https://rankings.ft.com/rankings/2948/masters-in-management-2023",
    ("masters-in-management", 2024): "https://rankings.ft.com/rankings/2961/masters-in-management-2024",
    ("masters-in-management", 2025): "https://rankings.ft.com/rankings/3004/masters-in-management-2025",
    ("masters-in-finance-pre-experience", 2023): "https://rankings.ft.com/rankings/2946/masters-in-finance-pre-experience-2023",
    ("masters-in-finance-pre-experience", 2024): "https://rankings.ft.com/rankings/2958/masters-in-finance-2024",
    ("masters-in-finance-pre-experience", 2025): "https://rankings.ft.com/rankings/3003/masters-in-finance-pre-experience-2025",
    ("masters-in-finance-post-experience", 2022): "https://rankings.ft.com/rankings/2870/masters-in-finance-post-experience-2022",
    ("masters-in-finance-post-experience", 2024): "https://rankings.ft.com/rankings/2959/masters-in-finance-post-experience-2024",
    ("masters-in-finance", 2024): "https://rankings.ft.com/rankings/2958/masters-in-finance-2024",
    ("executive-education-open", 2024): "https://rankings.ft.com/rankings/2956/executive-education-open-2024",
    ("executive-education-open", 2025): "https://rankings.ft.com/rankings/3001/executive-education-open-2025",
    ("executive-education-custom", 2023): "https://rankings.ft.com/rankings/2945/executive-education-custom-2023",
    ("executive-education-custom", 2024): "https://rankings.ft.com/rankings/2955/executive-education-custom-2024",
    ("executive-education-custom", 2025): "https://rankings.ft.com/rankings/3000/executive-education-custom-2025",
    ("online-mba", 2024): "https://rankings.ft.com/rankings/2953/online-mba-2024",
    ("online-mba", 2025): "https://rankings.ft.com/rankings/2998/online-mba-2025",
    ("mba", 2025): "https://rankings.ft.com/rankings/2997/mba-2025",
    ("emba", 2025): "https://rankings.ft.com/rankings/3005/emba-2025",
    ("european-business-school-rankings", 2021): "https://rankings.ft.com/rankings/2869/european-business-school-rankings-2021",
    ("european-business-school-rankings", 2022): "https://rankings.ft.com/rankings/2943/european-business-school-rankings-2022",
    ("european-business-school-rankings", 2023): "https://rankings.ft.com/rankings/2954/european-business-school-rankings-2023",
    ("european-business-school-rankings", 2024): "https://rankings.ft.com/rankings/2999/european-business-school-rankings-2024",
    ("european-business-school-rankings", 2025): "https://rankings.ft.com/rankings/3042/european-business-school-rankings-2025",
}


@dataclass
class FileMeta:
    master_type: str
    category: str
    year: int
    source_url: str


def _clean_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _normalize_header(value: str) -> str:
    lowered = value.lower()
    cleaned = re.sub(r"[^a-z0-9#]+", " ", lowered)
    return _clean_text(cleaned)


def _match_header(header: str) -> Optional[str]:
    normalized = _normalize_header(header)
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = _normalize_header(alias)
            if normalized == alias_norm or normalized.startswith(alias_norm):
                return canonical
    return None


def _parse_rank(raw: Optional[str]) -> Optional[int]:
    if raw is None:
        return None
    match = re.search(r"\d+(?:[.,]\d+)?", str(raw))
    if not match:
        return None
    number_str = match.group(0).replace(",", ".")
    try:
        return int(float(number_str))
    except ValueError:
        return None


def _parse_score(raw: Optional[str]) -> Optional[float]:
    if raw is None:
        return None
    cleaned = str(raw).replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _infer_meta_from_filename(path: Path) -> FileMeta:
    stem = path.stem.lower()
    year_match = re.search(r"(20\d{2})", stem)
    if not year_match:
        raise ValueError(f"Cannot infer year from filename: {path.name}")
    year = int(year_match.group(1))

    matched_pattern: Optional[Tuple[str, str, str]] = None
    for pattern, master_type, category in FILENAME_PATTERNS:
        if pattern in stem:
            matched_pattern = (pattern, master_type, category)
            break
    if matched_pattern is None:
        raise ValueError(f"Cannot infer master_type/category from filename: {path.name}")

    pattern, master_type, category = matched_pattern
    source_url = URL_MAP.get((pattern, year), "https://rankings.ft.com")

    return FileMeta(
        master_type=master_type,
        category=category,
        year=year,
        source_url=source_url,
    )


def _header_row(ws) -> Tuple[int, List[str]]:
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [cell for cell in row if cell not in (None, "")]
        if len(values) >= 2:
            return idx, [str(cell) for cell in row]
    raise ValueError("No header row found in workbook")


def _build_header_map(headers: List[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, header in enumerate(headers):
        canonical = _match_header(str(header))
        if canonical:
            mapping[idx] = canonical
    return mapping


def _extract_cell_value(cell: Cell) -> Optional[str]:
    if cell.value is None:
        return None
    return str(cell.value)


def _extract_link(cell: Cell) -> Optional[str]:
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None


def _parse_rows(ws, header_row_idx: int, header_map: Dict[int, str]) -> List[RankingEntry]:
    if "rank" not in header_map.values() or "school_name" not in header_map.values():
        raise ValueError("Table must contain rank and school columns")

    entries: List[RankingEntry] = []
    has_link_column = "link" in header_map.values()
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        values_present = any(cell.value not in (None, "") for cell in row)
        if not values_present:
            continue

        normalized: Dict[str, Optional[str]] = {key: None for key in header_map.keys()}
        school_link: Optional[str] = None

        for idx, cell in enumerate(row):
            canonical = header_map.get(idx)
            if canonical is None:
                continue
            if canonical == "link":
                normalized[canonical] = _extract_link(cell) or _extract_cell_value(cell)
            else:
                if canonical == "school_name":
                    school_link = _extract_link(cell)
                normalized[canonical] = _extract_cell_value(cell)

        rank_int = _parse_rank(normalized.get("rank"))
        school_value = normalized.get("school_name")
        if rank_int is None or not school_value:
            continue

        link_value = normalized.get("link") or (school_link if not has_link_column else None)
        entry = RankingEntry(
            rank=rank_int,
            school_name=_clean_text(school_value),
            program_name=_clean_text(normalized["program_name"]) if normalized.get("program_name") else None,
            country=_clean_text(normalized["country"]) if normalized.get("country") else None,
            city=_clean_text(normalized["city"]) if normalized.get("city") else None,
            score=_parse_score(normalized.get("score")),
            notes=None,
            link=link_value,
            metadata={},
        )
        entries.append(entry)

    if not entries:
        raise ValueError("No entries parsed from workbook")
    return entries


def convert_file(path: Path, output_dir: Path) -> Path:
    meta = _infer_meta_from_filename(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row_idx, headers = _header_row(ws)
    header_map = _build_header_map(headers)
    entries = _parse_rows(ws, header_row_idx, header_map)

    payload = LeaderboardPayload(
        master_type=meta.master_type,
        source="Financial Times",
        category=meta.category,
        year=meta.year,
        source_url=meta.source_url,
        region=None,
        entries=entries,
    )
    payload.validate()

    slug = _slugify(f"ft-{meta.category}-{meta.year}")
    output_path = output_dir / f"{slug}.json"
    output_path.write_text(json.dumps(payload.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8")
    return output_path


def _slugify(value: str) -> str:
    normalized = (
        value.lower()
        .replace("&", "and")
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")
    return normalized or "ranking"


def discover_files(input_dir: Path) -> Iterable[Path]:
    for path in sorted(input_dir.glob("*.xlsx")):
        if path.is_file():
            yield path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert FT Excel rankings to normalized JSON.")
    parser.add_argument("--input-dir", default="data/financial_times", help="Directory containing FT Excel exports.")
    parser.add_argument("--output-dir", default="data/rankings", help="Directory to write JSON payloads.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    files = list(discover_files(input_dir))
    if not files:
        raise SystemExit(f"No .xlsx files found in {input_dir}")

    for file_path in files:
        output_path = convert_file(file_path, output_dir)
        print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()

